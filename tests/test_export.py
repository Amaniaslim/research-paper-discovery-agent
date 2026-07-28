from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

import export_utils
import paper_research_agent as pipeline
import rag_answer


def test_markdown_export(tmp_path) -> None:
    review = pipeline.build_review(
        "security risks of agentic AI systems",
        limit=2,
        live=False,
    )

    markdown = pipeline.export_review(review, tmp_path / "review.md")

    assert "Literature Review" in markdown
    assert "security risks of agentic AI systems" in markdown
    assert (tmp_path / "review.md").exists()


def test_answer_fallback_without_api_key(monkeypatch) -> None:
    monkeypatch.setenv("SAIA_API_KEY", "")
    chunks = [
        {
            "pdf_name": "paper.pdf",
            "page_number": 2,
            "chunk_id": "p2-c1",
            "text": "The document discusses prompt injection and tool misuse risks.",
        }
    ]

    result = rag_answer.generate_answer("Which risks are discussed?", chunks)

    assert result["mode"] == "fallback"
    assert "Auf Basis der gefundenen PDF-Quellen" in result["answer"]
    assert "[paper.pdf, page 2, chunk p2-c1]" in result["answer"]
    assert "keine vollstaendige Analyse des gesamten Dokuments" in result["answer"]


def test_excel_source_export_contains_metadata() -> None:
    rows = export_utils.build_rows(
        question="What controls are discussed?",
        answer="Access controls are discussed.",
        chunks=[
            {
                "pdf_name": "paper.pdf",
                "page_number": 4,
                "chunk_id": "p4-c1",
                "section": "Methods",
                "extraction_method": "pdf_text",
                "retrieval_score": 3.5,
                "text": "Access controls for vector databases.",
            }
        ],
        timestamp="2026-07-28 12:00 UTC",
        embedding_model="local",
        ocr_mode="disabled",
        page_range="1-10",
    )

    workbook_bytes = export_utils.to_excel(rows)

    assert workbook_bytes is not None
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook["Sources"]
    assert worksheet.cell(row=2, column=3).value == "paper.pdf"
    assert worksheet.cell(row=2, column=5).value == "p4-c1"
